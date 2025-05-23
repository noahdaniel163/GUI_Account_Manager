import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from utils.logger import Logger
from openpyxl import load_workbook

logger = Logger('excel_tab')

class ExcelTab:
    def __init__(self, parent, api_handler, simo_converter):
        self.parent = parent
        self.api_handler = api_handler
        self.simo_converter = simo_converter
        
        # Frame chính
        self.excel_tab = ttk.Frame(parent)
        
        # Top frame cho tab
        self.excel_top_frame = ttk.Frame(self.excel_tab)
        self.excel_top_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.create_excel_section()
        
    def get_tab_frame(self):
        return self.excel_tab
        
    def create_excel_section(self):
        """Tạo giao diện cho tab xử lý Excel"""
        # File selection frame
        file_frame = ttk.LabelFrame(self.excel_top_frame, text="Chọn file Excel", padding="10")
        file_frame.pack(fill=tk.X, padx=5, pady=5)
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.entry_file = ttk.Entry(file_input_frame, width=100)
        self.entry_file.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        browse_btn = ttk.Button(file_input_frame, text="Browse", command=self.select_file, style='Primary.TButton')
        browse_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Service selection frame
        service_frame = ttk.LabelFrame(self.excel_top_frame, text="Chọn loại dịch vụ", padding="10")
        service_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.service_var = tk.StringVar(value="simo_001")
        services = ["simo_001", "simo_002", "simo_003", "simo_004", "simo_011", "simo_012"]
        service_combo = ttk.Combobox(service_frame, textvariable=self.service_var, 
                                   values=services, state='readonly', width=30)
        service_combo.pack(padx=5, pady=5)
        
        # Buttons frame
        button_frame = ttk.Frame(self.excel_top_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        convert_btn = ttk.Button(button_frame, text="Chuyển đổi Excel sang JSON", 
                               command=self.convert_excel, style='Primary.TButton')
        convert_btn.pack(side=tk.LEFT, padx=5)
        
        send_btn = ttk.Button(button_frame, text="Gửi dữ liệu", 
                            command=self.send_data, style='Success.TButton')
        send_btn.pack(side=tk.LEFT, padx=5)
        
        save_btn = ttk.Button(button_frame, text="Lưu JSON", 
                            command=self.save_json, style='Info.TButton')
        save_btn.pack(side=tk.LEFT, padx=5)
        
        # Paned window để chia màn hình thành 2 phần
        self.paned = ttk.PanedWindow(self.excel_tab, orient=tk.HORIZONTAL)
        self.paned.pack(fill=tk.BOTH, expand=True)
        
        # Excel display frame (left side)
        excel_frame = ttk.LabelFrame(self.paned, text="Nội dung Excel", padding="10")
        self.paned.add(excel_frame, weight=1)
        
        # Tạo frame chứa Treeview và scrollbar cho Excel
        excel_container = ttk.Frame(excel_frame)
        excel_container.pack(fill=tk.BOTH, expand=True)
        
        # Create Excel Treeview with scrollbars
        self.excel_tree = ttk.Treeview(excel_container, show="headings", style='Treeview')
        
        # Tạo scrollbars cho Excel
        excel_scroll_y = ttk.Scrollbar(excel_container, orient=tk.VERTICAL, 
                                     command=self.excel_tree.yview)
        excel_scroll_x = ttk.Scrollbar(excel_frame, orient=tk.HORIZONTAL, 
                                     command=self.excel_tree.xview)
        
        # Cấu hình scrollbar cho Excel
        self.excel_tree.configure(yscrollcommand=excel_scroll_y.set,
                                xscrollcommand=excel_scroll_x.set)
        
        # Pack Excel components
        excel_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.excel_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        excel_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # JSON display frame (right side)
        json_frame = ttk.LabelFrame(self.paned, text="Nội dung JSON", padding="10")
        self.paned.add(json_frame, weight=1)
        
        # Tạo frame chứa Text widget và scrollbar cho JSON
        json_container = ttk.Frame(json_frame)
        json_container.pack(fill=tk.BOTH, expand=True)
        
        # Create JSON Text widget with scrollbars
        self.json_text = tk.Text(json_container, wrap=tk.NONE, font=('Consolas', 10),
                               bg='#f8f9fa', fg='#212529', padx=10, pady=10)
        
        # Tạo scrollbars cho JSON
        json_scroll_y = ttk.Scrollbar(json_container, orient=tk.VERTICAL,
                                    command=self.json_text.yview)
        json_scroll_x = ttk.Scrollbar(json_frame, orient=tk.HORIZONTAL,
                                    command=self.json_text.xview)
        
        # Cấu hình scrollbar cho JSON
        self.json_text.configure(yscrollcommand=json_scroll_y.set,
                               xscrollcommand=json_scroll_x.set)
        
        # Pack JSON components
        json_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.json_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        json_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, file_path)
            self.load_excel_data(file_path)
            
    def load_excel_data(self, file_path):
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Xóa dữ liệu cũ trong Treeview
            self.excel_tree.delete(*self.excel_tree.get_children())
            
            # Lấy headers từ dòng đầu tiên
            headers = [str(cell.value).strip() for cell in ws[1]]
            self.excel_tree["columns"] = headers
            
            # Định dạng các cột
            for header in headers:
                self.excel_tree.heading(header, text=header)
                self.excel_tree.column(header, width=100)
            
            # Thêm dữ liệu vào Treeview
            for row in ws.iter_rows(min_row=2):
                values = [str(cell.value) if cell.value is not None else "" for cell in row]
                self.excel_tree.insert("", tk.END, values=values)
                
            logger.info(f"Đã tải dữ liệu từ file: {file_path}")
            messagebox.showinfo("Thành công", "Đã tải dữ liệu Excel!")
            
        except Exception as e:
            logger.error(f"Lỗi khi đọc file Excel: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể đọc file Excel: {str(e)}")
    
    def convert_excel(self):
        file_path = self.entry_file.get()
        service_type = self.service_var.get()
        
        if not file_path or not os.path.exists(file_path):
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel hợp lệ!")
            return
            
        try:
            # Chuyển đổi Excel sang JSON
            data = self.simo_converter.convert_excel_to_json(file_path, service_type)
            
            # Hiển thị JSON
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            self.json_text.delete(1.0, tk.END)
            self.json_text.insert(tk.END, json_str)
            
            logger.info(f"Đã chuyển đổi dữ liệu từ {file_path} sang JSON")
            messagebox.showinfo("Thành công", "Đã chuyển đổi dữ liệu thành công!")
            
        except Exception as e:
            logger.error(f"Lỗi khi chuyển đổi dữ liệu: {str(e)}")
            messagebox.showerror("Lỗi", f"Lỗi khi chuyển đổi dữ liệu: {str(e)}")
    
    def save_json(self):
        json_content = self.json_text.get(1.0, tk.END).strip()
        if not json_content:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu JSON để lưu!")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                initialfile=f"payload_{self.service_var.get()}.json"
            )
            if file_path:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(json_content)
                logger.info(f"Đã lưu JSON vào file: {file_path}")
                messagebox.showinfo("Thành công", f"Đã lưu file JSON tại: {file_path}")
        except Exception as e:
            logger.error(f"Lỗi khi lưu file JSON: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể lưu file: {str(e)}")
    
    def send_data(self):
        try:
            json_content = self.json_text.get(1.0, tk.END).strip()
            if not json_content:
                messagebox.showwarning("Cảnh báo", "Không có dữ liệu JSON để gửi!")
                return
                
            service_type = self.service_var.get()
            payload = json.loads(json_content)
            
            # Gửi dữ liệu qua API
            result = self.api_handler.send_data(service_type, payload)
            
            if result:
                logger.info(f"Gửi dữ liệu {service_type} thành công")
                messagebox.showinfo("Thành công", 
                                  f"Đã gửi dữ liệu {service_type} thành công!\n\nKết quả: {result}")
            
        except Exception as e:
            logger.error(f"Lỗi khi gửi dữ liệu: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể gửi dữ liệu: {str(e)}")