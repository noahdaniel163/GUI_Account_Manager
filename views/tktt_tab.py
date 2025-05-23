import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from datetime import datetime
from utils.logger import Logger
from openpyxl import Workbook
from views.detail_dialog import DetailDialog
from views.preview_dialog import PreviewDialog

logger = Logger('tktt_tab')

class TKTTTab:
    def __init__(self, parent, db_handler, api_handler, simo_converter):
        self.parent = parent
        self.db_handler = db_handler
        self.api_handler = api_handler
        self.simo_converter = simo_converter
        
        # Frame chính
        self.tktt_ca_nhan_tab = ttk.Frame(parent)
        
        # Top frame cho tab
        self.tktt_ca_nhan_top_frame = ttk.Frame(self.tktt_ca_nhan_tab)
        self.tktt_ca_nhan_top_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Biến cho phân trang
        self.current_page = 1
        self.rows_per_page = 100
        self.total_records = 0
        self.total_pages = 0
        
        # Tạo giao diện
        self.create_tktt_section()
        
        # Add pagination frame
        self.pagination_frame = ttk.Frame(self.tktt_ca_nhan_top_frame)
        self.pagination_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(self.pagination_frame, text="<<", command=self.prev_page).pack(side=tk.LEFT, padx=2)
        self.page_label = ttk.Label(self.pagination_frame, text="Page 1")
        self.page_label.pack(side=tk.LEFT, padx=5)
        ttk.Button(self.pagination_frame, text=">>", command=self.next_page).pack(side=tk.LEFT, padx=2)
        
        self.total_label = ttk.Label(self.pagination_frame, text="Total: 0 records")
        self.total_label.pack(side=tk.LEFT, padx=20)
        
    def get_tab_frame(self):
        return self.tktt_ca_nhan_tab
        
    def create_tktt_section(self):
        """Tạo giao diện cho tab dữ liệu TKTT Cá nhân"""
        # Frame điều khiển
        control_frame = ttk.LabelFrame(self.tktt_ca_nhan_top_frame, text="Thao tác", padding="10")
        control_frame.pack(fill=tk.X, padx=5, pady=5)

        # Frame chứa các nút điều khiển
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        # Frame chọn mã SIMO cho xuất JSON
        simo_frame = ttk.Frame(control_frame)
        simo_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(simo_frame, text="Mã SIMO:").pack(side=tk.LEFT, padx=5)
        self.tktt_simo_var = tk.StringVar(value="simo_001")
        simo_codes = ["simo_001", "simo_002", "simo_003", "simo_004", "simo_011", "simo_012"]
        simo_combo = ttk.Combobox(simo_frame, textvariable=self.tktt_simo_var, 
                               values=simo_codes, state='readonly', width=15)
        simo_combo.pack(side=tk.LEFT, padx=5)
        
        # Thêm ghi chú về việc xuất JSON không ảnh hưởng đến Database
        json_note = ttk.Label(simo_frame, text="(Lưu ý: Dữ liệu xuất JSON chỉ để gửi đi, không làm thay đổi dữ liệu gốc)", 
                           font=('Segoe UI', 9, 'italic'), foreground='#555555')
        json_note.pack(side=tk.LEFT, padx=5)

        # Các nút điều khiển chính
        fetch_btn = ttk.Button(button_frame, text="Tải dữ liệu", 
                           command=self.fetch_tktt_data, style='Primary.TButton')
        fetch_btn.pack(side=tk.LEFT, padx=5)

        refresh_btn = ttk.Button(button_frame, text="Làm mới", 
                             command=self.refresh_tktt_data, style='Info.TButton')
        refresh_btn.pack(side=tk.LEFT, padx=5)

        export_excel_btn = ttk.Button(button_frame, text="Xuất Excel", 
                                  command=self.export_tktt_to_excel, style='Success.TButton')
        export_excel_btn.pack(side=tk.LEFT, padx=5)

        export_json_btn = ttk.Button(button_frame, text="Xuất JSON", 
                                 command=self.export_selected_to_simo_json, style='Primary.TButton')
        export_json_btn.pack(side=tk.LEFT, padx=5)

        # Frame tìm kiếm
        search_frame = ttk.LabelFrame(self.tktt_ca_nhan_top_frame, text="Tìm kiếm", padding="10")
        search_frame.pack(fill=tk.X, padx=5, pady=5)

        # Frame cho các trường tìm kiếm cơ bản
        basic_search_frame = ttk.Frame(search_frame)
        basic_search_frame.pack(fill=tk.X, padx=5, pady=5)

        # CIF/SoID search
        ttk.Label(basic_search_frame, text="CIF/SoID:").grid(row=0, column=0, padx=5, pady=5)
        self.search_cif_entry = ttk.Entry(basic_search_frame, width=20)
        self.search_cif_entry.grid(row=0, column=1, padx=5, pady=5)

        # Tên khách hàng search
        ttk.Label(basic_search_frame, text="Tên KH:").grid(row=0, column=2, padx=5, pady=5)
        self.search_name_entry = ttk.Entry(basic_search_frame, width=30)
        self.search_name_entry.grid(row=0, column=3, padx=5, pady=5)

        # Frame cho các nút tìm kiếm
        search_button_frame = ttk.Frame(search_frame)
        search_button_frame.pack(fill=tk.X, padx=5, pady=5)

        # Nút tìm kiếm và xóa tìm kiếm
        search_btn = ttk.Button(search_button_frame, text="Tìm kiếm",
                            command=self.search_tktt_data, style='Primary.TButton')
        search_btn.pack(side=tk.LEFT, padx=5)

        clear_btn = ttk.Button(search_button_frame, text="Xóa tìm kiếm",
                           command=self.clear_search, style='Info.TButton')
        clear_btn.pack(side=tk.LEFT, padx=5)

        # Frame hiển thị dữ liệu
        self.tktt_display_frame = ttk.Frame(self.tktt_ca_nhan_tab)
        self.tktt_display_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Tạo Treeview với scrollbar
        self.tktt_tree = ttk.Treeview(self.tktt_display_frame, show="headings", 
                                   selectmode='extended', style='Treeview')

        # Tạo scrollbar
        tktt_scroll_y = ttk.Scrollbar(self.tktt_display_frame, orient=tk.VERTICAL,
                                   command=self.tktt_tree.yview)
        tktt_scroll_x = ttk.Scrollbar(self.tktt_display_frame, orient=tk.HORIZONTAL,
                                   command=self.tktt_tree.xview)

        # Cấu hình scrollbar
        self.tktt_tree.configure(yscrollcommand=tktt_scroll_y.set,
                              xscrollcommand=tktt_scroll_x.set)

        # Pack các components
        tktt_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        tktt_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tktt_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bind double click event để xem chi tiết
        self.tktt_tree.bind("<Double-1>", self.show_detail_dialog)
        
    def refresh_tktt_data(self):
        """Làm mới dữ liệu TKTT"""
        self.tktt_tree.delete(*self.tktt_tree.get_children())
        self.fetch_tktt_data()

    def fetch_tktt_data(self):
        """Đọc và hiển thị dữ liệu từ bảng TKTT với phân trang"""
        try:
            # Get total records first
            count_query = "SELECT COUNT(*) FROM TKTT WHERE LoaiKhachHang = N'Ca Nhan'"
            total = self.db_handler.execute_query(count_query, fetchall=False)[0]
            self.total_records = total
            self.total_pages = (total + self.rows_per_page - 1) // self.rows_per_page
            
            # Calculate offset
            offset = (self.current_page - 1) * self.rows_per_page
            
            query = f"""
                SELECT 
                    CAST(Cif AS VARCHAR(36)) AS Cif,
                    CAST(Soid AS VARCHAR(15)) AS Soid,
                    CAST(LoaiD AS INT) AS LoaiD,
                    CAST(TenKhachHang AS NVARCHAR(150)) AS TenKhachHang,
                    CONVERT(VARCHAR(10), NgaySinh, 103) AS NgaySinh,
                    CAST(GioiTinh AS INT) AS GioiTinh,
                    CAST(MaSoThue AS VARCHAR(50)) AS MaSoThue,
                    CAST(SoDienThoaiDangKyDichVu AS VARCHAR(15)) AS SoDienThoaiDangKyDichVu,
                    CAST(DiaChi AS NVARCHAR(300)) AS DiaChi,
                    CAST(DiaChiKiemSoatTruyCap AS NVARCHAR(60)) AS DiaChiKiemSoatTruyCap,
                    CAST(MaSoNhanDangThietBiDong AS VARCHAR(36)) AS MaSoNhanDangThietBiDong,
                    CAST(SoTaiKhoan AS VARCHAR(50)) AS SoTaiKhoan,
                    CAST(LoaiTaiKhoan AS INT) AS LoaiTaiKhoan,
                    CAST(TrangThaiHoatDongTaiKhoan AS INT) AS TrangThaiHoatDongTaiKhoan,
                    CONVERT(VARCHAR(10), NgayMoTaiKhoan, 103) AS NgayMoTaiKhoan,
                    CAST(PhuongThucMoTaiKhoan AS INT) AS PhuongThucMoTaiKhoan,
                    CONVERT(VARCHAR(10), NgayXacThucTaiQuay, 103) AS NgayXacThucTaiQuay,
                    CAST(QuocTich AS NVARCHAR(36)) AS QuocTich,
                    CAST(LoaiKhachHang AS NVARCHAR(50)) AS LoaiKhachHang,
                    CAST(GhiChu AS NVARCHAR(500)) AS GhiChu,
                    CONVERT(VARCHAR(10), UpdateDate, 103) AS UpdateDate,
                    CAST(NghiNgo AS INT) AS NghiNgo
                FROM TKTT 
                WHERE LoaiKhachHang = N'Ca Nhan'
                ORDER BY UpdateDate DESC
                OFFSET ? ROWS
                FETCH NEXT ? ROWS ONLY
            """
            
            rows = self.db_handler.execute_query(query, params=(offset, self.rows_per_page))
            
            # Clear existing data
            self.tktt_tree.delete(*self.tktt_tree.get_children())
            
            # Get column names
            columns = [column[0] for column in rows[0].cursor_description] if rows else []
            
            # Update Treeview columns
            self.tktt_tree["columns"] = columns
            for col in columns:
                self.tktt_tree.heading(col, text=col)
                self.tktt_tree.column(col, width=100)
            
            # Add data to Treeview
            for row in rows:
                values = [str(val) if val is not None else "" for val in row]
                self.tktt_tree.insert("", tk.END, values=values)
            
            # Update pagination info
            self.update_pagination_info()
            
            logger.info(f"Đã đọc {len(rows)} bản ghi từ bảng TKTT (Trang {self.current_page}/{self.total_pages})")
            
        except Exception as e:
            logger.error(f"Lỗi khi đọc dữ liệu TKTT: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể đọc dữ liệu: {str(e)}")

    def update_pagination_info(self):
        """Cập nhật thông tin phân trang"""
        self.page_label.config(text=f"Trang {self.current_page}/{self.total_pages}")
        self.total_label.config(text=f"Tổng số: {self.total_records:,} bản ghi")

    def next_page(self):
        """Chuyển đến trang tiếp theo"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.fetch_tktt_data()

    def prev_page(self):
        """Quay lại trang trước"""
        if self.current_page > 1:
            self.current_page -= 1
            self.fetch_tktt_data()

    def search_tktt_data(self):
        """Tìm kiếm dữ liệu TKTT theo các điều kiện với phân trang"""
        try:
            conditions = []
            params = []
            
            # Reset pagination
            self.current_page = 1
            
            # Base query với điều kiện lọc khách hàng cá nhân
            base_query = """
                SELECT 
                    CAST(Cif AS VARCHAR(36)) AS Cif,
                    CAST(Soid AS VARCHAR(15)) AS Soid,
                    CAST(LoaiD AS INT) AS LoaiD,
                    CAST(TenKhachHang AS NVARCHAR(150)) AS TenKhachHang,
                    CONVERT(VARCHAR(10), NgaySinh, 103) AS NgaySinh,
                    CAST(GioiTinh AS INT) AS GioiTinh,
                    CAST(MaSoThue AS VARCHAR(50)) AS MaSoThue,
                    CAST(SoDienThoaiDangKyDichVu AS VARCHAR(15)) AS SoDienThoaiDangKyDichVu,
                    CAST(DiaChi AS NVARCHAR(300)) AS DiaChi,
                    CAST(DiaChiKiemSoatTruyCap AS NVARCHAR(60)) AS DiaChiKiemSoatTruyCap,
                    CAST(MaSoNhanDangThietBiDong AS VARCHAR(36)) AS MaSoNhanDangThietBiDong,
                    CAST(SoTaiKhoan AS VARCHAR(50)) AS SoTaiKhoan,
                    CAST(LoaiTaiKhoan AS INT) AS LoaiTaiKhoan,
                    CAST(TrangThaiHoatDongTaiKhoan AS INT) AS TrangThaiHoatDongTaiKhoan,
                    CONVERT(VARCHAR(10), NgayMoTaiKhoan, 103) AS NgayMoTaiKhoan,
                    CAST(PhuongThucMoTaiKhoan AS INT) AS PhuongThucMoTaiKhoan,
                    CONVERT(VARCHAR(10), NgayXacThucTaiQuay, 103) AS NgayXacThucTaiQuay,
                    CAST(QuocTich AS NVARCHAR(36)) AS QuocTich,
                    CAST(LoaiKhachHang AS NVARCHAR(50)) AS LoaiKhachHang,
                    CAST(GhiChu AS NVARCHAR(500)) AS GhiChu,
                    CONVERT(VARCHAR(10), UpdateDate, 103) AS UpdateDate,
                    CAST(NghiNgo AS INT) AS NghiNgo
                FROM TKTT 
                WHERE LoaiKhachHang = N'Ca Nhan'
            """

            # Tìm theo CIF hoặc SoID
            if self.search_cif_entry.get().strip():
                search_term = self.search_cif_entry.get().strip()
                conditions.append("(CAST(Cif AS VARCHAR(36)) LIKE ? OR CAST(Soid AS VARCHAR(15)) LIKE ?)")
                params.extend([f"%{search_term}%", f"%{search_term}%"])
            
            # Tìm theo tên khách hàng
            if self.search_name_entry.get().strip():
                conditions.append("TenKhachHang LIKE ?")
                params.append(f"%{self.search_name_entry.get().strip()}%")
            
            where_clause = ""
            if conditions:
                where_clause = " AND " + " AND ".join(conditions)
            
            # Get total records for search
            count_query = f"SELECT COUNT(*) FROM TKTT WHERE LoaiKhachHang = N'Ca Nhan'{where_clause}"
            total = self.db_handler.execute_query(count_query, params=params, fetchall=False)[0]
            self.total_records = total
            self.total_pages = (total + self.rows_per_page - 1) // self.rows_per_page
            
            # Calculate offset for pagination
            offset = (self.current_page - 1) * self.rows_per_page
            
            # Final query with pagination
            query = f"{base_query}{where_clause} ORDER BY UpdateDate DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
            params.extend([offset, self.rows_per_page])
            
            rows = self.db_handler.execute_query(query, params=params)
            
            # Clear existing data
            self.tktt_tree.delete(*self.tktt_tree.get_children())
            
            # Add data to Treeview
            for row in rows:
                values = [str(val) if val is not None else "" for val in row]
                self.tktt_tree.insert("", tk.END, values=values)
            
            # Update pagination info
            self.update_pagination_info()
            
            # Display search results
            conditions_text = []
            if self.search_cif_entry.get().strip():
                conditions_text.append(f"CIF/SoID: {self.search_cif_entry.get().strip()}")
            if self.search_name_entry.get().strip():
                conditions_text.append(f"Tên KH: {self.search_name_entry.get().strip()}")
            
            result_message = f"Tìm thấy {total:,} bản ghi"
            if conditions_text:
                result_message += f"\nĐiều kiện tìm kiếm: {', '.join(conditions_text)}"
            
            messagebox.showinfo("Kết quả tìm kiếm", result_message)
            
        except Exception as e:
            logger.error(f"Lỗi khi tìm kiếm: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể tìm kiếm: {str(e)}")

    def clear_search(self):
        """Xóa các điều kiện tìm kiếm"""
        self.search_cif_entry.delete(0, tk.END)
        self.search_name_entry.delete(0, tk.END)
        self.fetch_tktt_data()

    def show_detail_dialog(self, event):
        """Hiển thị dialog chi tiết khi double click vào một dòng"""
        # Lấy item được chọn
        selected_item = self.tktt_tree.selection()
        if not selected_item:
            return
            
        # Lấy dữ liệu của dòng được chọn
        values = self.tktt_tree.item(selected_item[0])["values"]
        
        # Tạo dictionary từ values và tên cột
        columns = self.tktt_tree["columns"]
        data_dict = dict(zip(columns, values))
        
        # Hiển thị dialog chi tiết
        DetailDialog(self.parent, data_dict)
        
    def export_tktt_to_excel(self):
        """Xuất dữ liệu TKTT ra file Excel"""
        try:
            # Lấy tên cột và dữ liệu từ Treeview
            columns = self.tktt_tree["columns"]
            data = []
            for item in self.tktt_tree.get_children():
                values = self.tktt_tree.item(item)["values"]
                data.append(values)

            if not data:
                messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
                return

            # Chọn vị trí lưu file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="TKTT_Data.xlsx"
            )
            
            if file_path:
                # Tạo workbook và worksheet mới
                wb = Workbook()
                ws = wb.active
                
                # Thêm headers
                for col, header in enumerate(columns, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Thêm dữ liệu
                for row, record in enumerate(data, 2):
                    for col, value in enumerate(record, 1):
                        ws.cell(row=row, column=col, value=value)
                
                # Lưu file
                wb.save(file_path)
                logger.info(f"Đã xuất dữ liệu TKTT ra file: {file_path}")
                messagebox.showinfo("Thành công", f"Đã xuất dữ liệu ra file:\n{file_path}")

        except Exception as e:
            logger.error(f"Lỗi khi xuất dữ liệu: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu: {str(e)}")
        
    def export_selected_to_simo_json(self):
        """Xuất dữ liệu được chọn thành JSON"""
        try:
            selected_items = self.tktt_tree.selection()
            if not selected_items:
                messagebox.showwarning(
                    "Chưa chọn dữ liệu", 
                    "Vui lòng chọn ít nhất một dòng dữ liệu bằng cách:\n" +
                    "1. Click chọn một dòng hoặc\n" +
                    "2. Giữ Ctrl + Click để chọn nhiều dòng hoặc\n" +
                    "3. Giữ Shift + Click để chọn một đoạn liên tiếp"
                )
                return

            service_type = self.tktt_simo_var.get()
            columns = self.tktt_tree["columns"]
            
            # Tạo danh sách chứa dữ liệu được chọn với định dạng đúng
            selected_data = []
            for item in selected_items:
                values = self.tktt_tree.item(item)["values"]
                row_data = {}
                for col, val in zip(columns, values):
                    formatted_value = self.format_tree_value(col, val)
                    if formatted_value is not None:
                        row_data[col] = formatted_value
                selected_data.append(row_data)
            
            # Kiểm tra nếu không có dữ liệu hợp lệ
            if not selected_data:
                logger.warning(f"Không có dữ liệu hợp lệ trong {len(selected_items)} dòng được chọn")
                messagebox.showwarning(
                    "Dữ liệu không hợp lệ", 
                    "Các dòng được chọn không có dữ liệu hợp lệ để xuất.\n" +
                    "Vui lòng kiểm tra lại dữ liệu!"
                )
                return
            
            # Chuyển đổi dữ liệu theo định dạng SIMO
            converted_data = []
            
            # Chuyển dữ liệu sang định dạng SIMO thông qua converter
            converted_data = self.simo_converter.convert_to_simo(selected_data, service_type)

            if not converted_data:
                logger.warning(f"Không có dữ liệu sau khi chuyển đổi sang định dạng {service_type}")
                messagebox.showwarning(
                    "Không có dữ liệu", 
                    f"Không có dữ liệu hợp lệ để xuất theo định dạng {service_type}.\n" +
                    "Vui lòng kiểm tra lại dữ liệu và định dạng SIMO được chọn!"
                )
                return
                
            # Kiểm tra tính hợp lệ của số điện thoại trước khi xuất JSON cho SIMO_001 và SIMO_004
            if service_type in ["simo_001", "simo_004"]:
                is_valid, warnings = self.simo_converter.validate_data_before_export(converted_data, service_type)
                if not is_valid:
                    warning_message = "\n".join(warnings)
                    result = messagebox.askokcancel(
                        "Cảnh báo - Số điện thoại không đúng chuẩn Việt Nam", 
                        f"Phát hiện các lỗi với số điện thoại:\n\n{warning_message}\n\n" +
                        "Bạn có muốn tiếp tục xuất JSON không?\n" +
                        "(Ấn OK để tiếp tục, Cancel để hủy)"
                    )
                    if not result:
                        return

            # Hiển thị dialog xem trước kết quả chuyển đổi
            PreviewDialog(self.parent, converted_data, service_type)

        except Exception as e:
            logger.error(f"Lỗi khi xuất dữ liệu sang JSON: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể xuất dữ liệu: {str(e)}")
            return None
            
    def format_tree_value(self, column, value):
        """Format giá trị từ Treeview, giữ nguyên giá trị gốc từ database"""
        if value in [None, "None", ""]:
            return None
            
        # Các trường số nguyên
        int_fields = ["LoaiD", "GioiTinh", "LoaiTaiKhoan", "TrangThaiHoatDongTaiKhoan", 
                     "PhuongThucMoTaiKhoan", "NghiNgo"]
        if column in int_fields:
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
                
        # Trả về giá trị gốc cho tất cả các trường khác
        return str(value).strip()