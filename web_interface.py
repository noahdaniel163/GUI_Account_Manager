from pywebio.platform.flask import webio_view
from pywebio import start_server
from pywebio.input import *
from pywebio.output import *
from pywebio.session import *
from flask import Flask, send_from_directory
import os
import json
import base64
from datetime import datetime
from utils.api_handler import APIHandler
from utils.db_handler import DatabaseHandler
from utils.logger import Logger
from openpyxl import load_workbook

# Khởi tạo logger
logger = Logger('web_interface')

# Khởi tạo API handler
api_handler = APIHandler()

# Khởi tạo Flask app
app = Flask(__name__)

def convert_excel_to_json(file_path, service_type):
    """Chuyển đổi Excel sang JSON"""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Lấy headers từ dòng đầu tiên
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        # Đọc dữ liệu từ dòng thứ 2
        raw_data = []
        for row in ws.iter_rows(min_row=2):
            row_dict = {}
            for header, cell in zip(headers, row):
                value = cell.value if cell.value is not None else ""
                row_dict[header] = value
            raw_data.append(row_dict)

        # Chuyển đổi dữ liệu theo định dạng tương ứng
        if service_type == "simo_001":
            return convert_to_simo_001(raw_data)
        elif service_type == "simo_002":
            return convert_to_simo_002(raw_data)
        elif service_type == "simo_003":
            return convert_to_simo_003(raw_data)
        elif service_type == "simo_004":
            return convert_to_simo_004(raw_data)
        elif service_type == "simo_011":
            return convert_to_simo_011(raw_data)
        elif service_type == "simo_012":
            return convert_to_simo_012(raw_data)
        else:
            raise ValueError(f"Không hỗ trợ chuyển đổi cho loại dịch vụ: {service_type}")
            
    except Exception as e:
        logger.error(f"Lỗi khi chuyển đổi Excel sang JSON: {str(e)}")
        raise

def convert_to_simo_001(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 001"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoID": str(row.get("SoID", "")),
            "LoaiID": int(row.get("LoaiID", 0)),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "NgaySinh": str(row.get("NgaySinh", "")),
            "GioiTinh": int(row.get("GioiTinh", 0)),
            "MaSoThue": str(row.get("MaSoThue", "")),
            "SoDienThoaiDangKyDichVu": str(row.get("SoDienThoaiDangKyDichVu", "")),
            "DiaChi": str(row.get("DiaChi", "")),
            "DiaChiKiemSoatTruyCap": str(row.get("DiaChiKiemSoatTruyCap", "")),
            "MaSoNhanDangThietBiDiDong": str(row.get("MaSoNhanDangThietBiDiDong", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "LoaiTaiKhoan": int(row.get("LoaiTaiKhoan", 0)),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NgayMoTaiKhoan": str(row.get("NgayMoTaiKhoan", "")),
            "PhuongThucMoTaiKhoan": int(row.get("PhuongThucMoTaiKhoan", 0)),
            "NgayXacThucTaiQuay": str(row.get("NgayXacThucTaiQuay", "")),
            "QuocTich": str(row.get("QuocTich", ""))
        }
        payload.append(record)
    return payload

def convert_to_simo_002(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 002"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NghiNgo": int(row.get("NghiNgo", 0)),
            "GhiChu": str(row.get("GhiChu", ""))[:500]
        }
        payload.append(record)
    return payload

def convert_to_simo_003(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 003"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NghiNgo": int(row.get("NghiNgo", 0)),
            "GhiChu": str(row.get("GhiChu", ""))[:500]
        }
        payload.append(record)
    return payload

def convert_to_simo_004(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 004"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoID": str(row.get("SoID", "")),
            "LoaiID": int(row.get("LoaiID", 0)),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "NgaySinh": str(row.get("NgaySinh", "")),
            "GioiTinh": int(row.get("GioiTinh", 0)),
            "MaSoThue": str(row.get("MaSoThue", "")),
            "SoDienThoaiDangKyDichVu": str(row.get("SoDienThoaiDangKyDichVu", "")),
            "DiaChi": str(row.get("DiaChi", "")),
            "DiaChiKiemSoatTruyCap": str(row.get("DiaChiKiemSoatTruyCap", "")),
            "MaSoNhanDangThietBiDiDong": str(row.get("MaSoNhanDangThietBiDiDong", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "LoaiTaiKhoan": int(row.get("LoaiTaiKhoan", 0)),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NgayMoTaiKhoan": str(row.get("NgayMoTaiKhoan", "")),
            "PhuongThucMoTaiKhoan": int(row.get("PhuongThucMoTaiKhoan", 0)),
            "NgayXacThucTaiQuay": str(row.get("NgayXacThucTaiQuay", "")),
            "GhiChu": str(row.get("GhiChu", ""))[:500],
            "QuocTich": str(row.get("QuocTich", ""))
        }
        payload.append(record)
    return payload

def convert_to_simo_011(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 011"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoID": str(row.get("SoID", "")),
            "LoaiID": int(row.get("LoaiID", 0)),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "NgaySinh": str(row.get("NgaySinh", "")),
            "GioiTinh": int(row.get("GioiTinh", 0)),
            "MaSoThue": str(row.get("MaSoThue", "")),
            "SoDienThoaiDangKyDichVu": str(row.get("SoDienThoaiDangKyDichVu", "")),
            "DiaChi": str(row.get("DiaChi", "")),
            "DiaChiKiemSoatTruyCap": str(row.get("DiaChiKiemSoatTruyCap", "")),
            "MaSoNhanDangThietBiDiDong": str(row.get("MaSoNhanDangThietBiDiDong", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "LoaiTaiKhoan": int(row.get("LoaiTaiKhoan", 0)),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NgayMoTaiKhoan": str(row.get("NgayMoTaiKhoan", "")),
            "PhuongThucMoTaiKhoan": int(row.get("PhuongThucMoTaiKhoan", 0)),
            "NgayXacThucTaiQuay": str(row.get("NgayXacThucTaiQuay", "")),
            "QuocTich": str(row.get("QuocTich", ""))
        }
        payload.append(record)
    return payload

def convert_to_simo_012(data):
    """Chuyển đổi dữ liệu sang định dạng SIMO 012"""
    payload = []
    for row in data:
        record = {
            "Cif": str(row.get("Cif", "")),
            "SoTaiKhoan": str(row.get("SoTaiKhoan", "")),
            "TenKhachHang": str(row.get("TenKhachHang", "")),
            "TrangThaiHoatDongTaiKhoan": int(row.get("TrangThaiHoatDongTaiKhoan", 0)),
            "NghiNgo": int(row.get("NghiNgo", 0)),
            "GhiChu": str(row.get("GhiChu", ""))[:500]
        }
        payload.append(record)
    return payload

def simo_web_app():
    """Giao diện web chính của ứng dụng"""
    put_markdown("# SIMO GUI Enhanced - Web Interface")
    
    # Tạo tabs
    tabs = ['Chuyển đổi Excel', 'Cấu hình API', 'Logs']
    
    def handle_excel_tab():
        """Xử lý tab chuyển đổi Excel"""
        put_markdown("## Chuyển đổi Excel sang JSON")
        
        # Chọn loại dịch vụ
        service_type = select("Chọn loại dịch vụ:", [
            {'label': 'SIMO 001', 'value': 'simo_001'},
            {'label': 'SIMO 002', 'value': 'simo_002'},
            {'label': 'SIMO 003', 'value': 'simo_003'},
            {'label': 'SIMO 004', 'value': 'simo_004'},
            {'label': 'SIMO 011', 'value': 'simo_011'},
            {'label': 'SIMO 012', 'value': 'simo_012'}
        ])
        
        # Upload file Excel
        file = file_upload("Chọn file Excel:", accept=".xlsx,.xls")
        if not file:
            return
        
        # Lưu file tạm thời
        temp_file = f"temp_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(file['content'])
        
        try:
            # Chuyển đổi Excel sang JSON
            with put_loading():
                data = convert_excel_to_json(temp_file, service_type)
            
            # Hiển thị JSON
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            
            # Tạo container cho JSON
            put_markdown("### Kết quả chuyển đổi:")
            put_code(json_str, language='json')
            
            # Tạo nút để tải xuống JSON
            put_button("Tải xuống JSON", onclick=lambda: download(f"payload_{service_type}.json", json_str))
            
            # Tạo nút để gửi dữ liệu
            put_button("Gửi dữ liệu", onclick=lambda: send_data(service_type, data))
            
            logger.info(f"Đã chuyển đổi dữ liệu từ {file['filename']} sang JSON")
            toast("Đã chuyển đổi dữ liệu thành công!")
            
        except Exception as e:
            logger.error(f"Lỗi khi chuyển đổi dữ liệu: {str(e)}")
            toast(f"Lỗi: {str(e)}", color='error')
        finally:
            # Xóa file tạm
            if os.path.exists(temp_file):
                os.remove(temp_file)
    
    def send_data(service_type, payload):
        """Gửi dữ liệu qua API"""
        try:
            with put_loading():
                result = api_handler.send_data(service_type, payload)
            
            if result:
                logger.info(f"Gửi dữ liệu {service_type} thành công")
                popup("Thành công", [
                    put_markdown(f"Đã gửi dữ liệu {service_type} thành công!"),
                    put_markdown(f"Kết quả: ```{json.dumps(result, ensure_ascii=False, indent=2)}```"),
                    put_button("Đóng", onclick=lambda: close_popup())
                ])
        except Exception as e:
            logger.error(f"Lỗi khi gửi dữ liệu: {str(e)}")
            popup("Lỗi", [
                put_markdown(f"Không thể gửi dữ liệu: {str(e)}"),
                put_button("Đóng", onclick=lambda: close_popup())
            ], size='large')
    
    def handle_config_tab():
        """Xử lý tab cấu hình API"""
        put_markdown("## Cấu hình API")
        
        # Khởi tạo DB handler
        db = DatabaseHandler()
        db.connect()
        
        # Lấy thông tin xác thực hiện tại
        credentials = db.get_api_credentials()
        username = credentials[0] if credentials else ""
        password = credentials[1] if credentials else ""
        consumer_key = credentials[2] if credentials else ""
        consumer_secret = credentials[3] if credentials else ""
        
        # Form cấu hình xác thực
        put_markdown("### Thông tin xác thực API")
        auth_data = input_group("Cấu hình xác thực", [
            input("Username:", name="username", value=username),
            input("Password:", name="password", type="password", value=password),
            input("Consumer Key:", name="consumer_key", value=consumer_key),
            input("Consumer Secret:", name="consumer_secret", value=consumer_secret)
        ])
        
        # Lưu thông tin xác thực
        if db.save_api_credentials(auth_data["username"], auth_data["password"], 
                                auth_data["consumer_key"], auth_data["consumer_secret"]):
            toast("Đã lưu thông tin xác thực API")
        else:
            toast("Lỗi khi lưu thông tin xác thực", color='error')
        
        # Form cấu hình endpoint
        put_markdown("### Cấu hình Endpoint")
        
        # Lấy danh sách endpoint hiện tại
        endpoints = {}
        for endpoint in ['token', 'simo_001', 'simo_002', 'simo_003', 'simo_004', 'simo_011', 'simo_012']:
            url = db.get_endpoint_url(endpoint)
            endpoints[endpoint] = url if url else ""
        
        # Hiển thị form cấu hình endpoint
        endpoint_data = input_group("Cấu hình endpoint", [
            input(f"URL {endpoint.upper()}:", name=endpoint, value=url)
            for endpoint, url in endpoints.items()
        ])
        
        # Lưu cấu hình endpoint
        success = True
        for endpoint, url in endpoint_data.items():
            if not db.save_endpoint_url(endpoint, url):
                success = False
        
        if success:
            toast("Đã lưu cấu hình endpoint")
        else:
            toast("Có lỗi khi lưu cấu hình endpoint", color='error')
        
        db.close()
    
    def handle_logs_tab():
        """Xử lý tab logs"""
        put_markdown("## Logs")
        
        # Lấy danh sách file log
        log_dir = 'logs'
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        log_files = [f for f in os.listdir(log_dir) if f.endswith('.log')]
        
        if not log_files:
            put_markdown("Không có file log nào.")
            return
        
        # Chọn file log
        selected_log = select("Chọn file log:", options=log_files)
        
        # Hiển thị nội dung log
        with open(os.path.join(log_dir, selected_log), 'r', encoding='utf-8') as f:
            log_content = f.read()
        
        put_markdown(f"### Nội dung file {selected_log}:")
        put_code(log_content)
        
        # Tạo nút để tải xuống log
        put_button("Tải xuống log", onclick=lambda: download(selected_log, log_content))
    
    # Tạo tabs
    with use_scope('main'):
        put_tabs([
            {'title': 'Chuyển đổi Excel', 'content': handle_excel_tab},
            {'title': 'Cấu hình API', 'content': handle_config_tab},
            {'title': 'Logs', 'content': handle_logs_tab}
        ])

# Đăng ký route cho ứng dụng web
app.add_url_rule('/tool', 'webio_view', webio_view(simo_web_app), methods=['GET', 'POST', 'OPTIONS'])

# Route cho trang chủ
@app.route('/')
def index():
    return '<h1>SIMO GUI Enhanced - Web Interface</h1><p>Truy cập <a href="/tool">công cụ</a> để sử dụng ứng dụng.</p>'

# Route cho các file tĩnh
@app.route('/static/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

if __name__ == '__main__':
    # Tạo thư mục logs nếu chưa tồn tại
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Chạy ứng dụng web
    app.run(host='0.0.0.0', port=8080, debug=True)