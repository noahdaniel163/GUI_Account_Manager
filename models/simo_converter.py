from datetime import datetime, timedelta
from utils.logger import Logger

logger = Logger('simo_converter')

class SimoConverter:
    @staticmethod
    def get_default_value(value, field_type, max_length=None):
        """
        Xử lý giá trị mặc định cho các trường dữ liệu
        :param value: Giá trị từ nguồn (Excel hoặc Database)
        :param field_type: Kiểu dữ liệu mong muốn ('str', 'int', 'date')
        :param max_length: Độ dài tối đa cho chuỗi
        :return: Giá trị đã được xử lý hoặc None nếu giá trị rỗng
        """
        try:
            # Xử lý giá trị None hoặc rỗng
            if value is None or (isinstance(value, str) and not value.strip()):
                return None
                    
            # Xử lý theo kiểu dữ liệu
            if field_type == 'str':
                result = str(value).strip()
                if not result:
                    return None
                if max_length:
                    result = result[:max_length]
                return result
                
            elif field_type == 'int':
                try:
                    return int(float(str(value).strip()))
                except (ValueError, TypeError):
                    return None
                    
            elif field_type == 'date':
                # Nếu giá trị đã đúng định dạng dd/mm/yyyy, trả về luôn
                if isinstance(value, str):
                    value = value.strip()
                    try:
                        datetime.strptime(value, "%d/%m/%Y")
                        return value
                    except ValueError:
                        pass
                        
                # Nếu là datetime object
                if isinstance(value, datetime):
                    return value.strftime("%d/%m/%Y")
                
                # Nếu là string, thử chuyển đổi từ các định dạng khác
                if isinstance(value, str):
                    date_formats = [
                        ("%Y-%m-%d", "dd/mm/yyyy"),
                        ("%d-%m-%Y", "dd/mm/yyyy"),
                        ("%Y/%m/%d", "dd/mm/yyyy"),
                        ("%d.%m.%Y", "dd/mm/yyyy"),
                        ("%Y.%m.%d", "dd/mm/yyyy"),
                        ("%d-%m-%y", "dd/mm/yyyy"),
                        ("%d/%m/%y", "dd/mm/yyyy")
                    ]
                    
                    for input_format, _ in date_formats:
                        try:
                            date_obj = datetime.strptime(value, input_format)
                            return date_obj.strftime("%d/%m/%Y")
                        except ValueError:
                            continue
                            
                    # Nếu là số từ Excel (số ngày kể từ 1900)
                    try:
                        if float(value) > 0:
                            excel_epoch = datetime(1899, 12, 30)
                            date_obj = excel_epoch + timedelta(days=float(value))
                            return date_obj.strftime("%d/%m/%Y")
                    except (ValueError, TypeError):
                        pass
                        
                return None
                        
            return value
            
        except Exception as e:
            logger.warning(f"Lỗi xử lý giá trị {value}: {str(e)}")
            return None

    @staticmethod
    def format_number_with_padding(value, field):
        """Format số với độ dài chuẩn theo quy định VN"""
        if value is None or str(value).strip() == "":
            return None

        # Làm sạch giá trị đầu vào - loại bỏ khoảng trắng và dấu gạch ngang
        clean_value = str(value).strip().replace(" ", "").replace("-", "")
        
        # Xử lý theo từng loại trường
        if field == "SoDienThoaiDangKyDichVu":
            # Xử lý số điện thoại theo chuẩn VN
            
            # Loại bỏ mã quốc gia +84 nếu có
            if clean_value.startswith("+84"):
                clean_value = "0" + clean_value[3:]
            # Loại bỏ mã quốc gia 84 nếu có
            elif clean_value.startswith("84") and len(clean_value) >= 10:
                clean_value = "0" + clean_value[2:]
                
            # Đảm bảo số điện thoại bắt đầu bằng số 0
            if not clean_value.startswith("0"):
                clean_value = "0" + clean_value
                
            # Kiểm tra đầu số hợp lệ (03x, 05x, 07x, 08x, 09x)
            valid_prefixes = ["03", "05", "07", "08", "09"]
            if len(clean_value) >= 2 and clean_value[:2] not in valid_prefixes:
                logger.warning(f"Số điện thoại {clean_value} không có đầu số hợp lệ theo chuẩn VN")
            
            # Đảm bảo đúng 10 chữ số
            if len(clean_value) > 10:
                clean_value = clean_value[:10]
                logger.warning(f"Số điện thoại đã được cắt ngắn thành 10 số: {clean_value}")
            elif len(clean_value) < 10:
                logger.warning(f"Số điện thoại {clean_value} có độ dài không đúng chuẩn 10 số")
                
            return clean_value
            
        elif field == "SoID":
            # Xử lý số ID (CMND/CCCD) theo chuẩn VN
            # CMND cũ: 9 số, CCCD mới: 12 số
            
            # Nếu là ký tự chữ + số (Hộ chiếu) thì giữ nguyên
            if any(c.isalpha() for c in clean_value):
                return clean_value
            
            # Nếu là số, thì đảm bảo đúng độ dài theo chuẩn
            if clean_value.isdigit():
                # Xử lý CMND (9 số)
                if len(clean_value) < 9:
                    # Nếu độ dài nhỏ hơn 9, thêm số 0 ở đầu
                    padded_value = clean_value.zfill(9)
                    logger.info(f"Số ID {clean_value} đã được chuẩn hóa thành {padded_value} (CMND 9 số)")
                    return padded_value
                elif len(clean_value) == 9:
                    # Đúng độ dài CMND
                    return clean_value
                # Xử lý CCCD (12 số)
                elif len(clean_value) < 12 and len(clean_value) > 9:
                    # Nếu độ dài từ 10-11, thêm số 0 ở đầu để đủ 12 số (CCCD)
                    padded_value = clean_value.zfill(12)
                    logger.info(f"Số ID {clean_value} đã được chuẩn hóa thành {padded_value} (CCCD 12 số)")
                    return padded_value
                elif len(clean_value) == 12:
                    # Đúng độ dài CCCD
                    return clean_value
                else:
                    # Độ dài lớn hơn 12, ghi log cảnh báo
                    logger.warning(f"Số ID {clean_value} có độ dài {len(clean_value)} không đúng chuẩn CMND (9 số) hoặc CCCD (12 số) của VN")
                    return clean_value
            
            return clean_value
            
        elif field == "Cif":
            # Giữ nguyên giá trị CIF, thường là định dạng đặc biệt của ngân hàng
            return clean_value
            
        elif field == "SoTaiKhoan":
            # Số tài khoản thường có độ dài cố định (thường 13-16 số)
            # Cần giữ nguyên mọi số 0 ở đầu
            if not clean_value:
                return None
                
            return clean_value
            
        else:
            # Các trường khác giữ nguyên
            return clean_value

    @staticmethod
    def convert_to_simo_001(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_001"""
        payload = []
        field_mapping = {
            "Cif": "Cif",
            "Soid": "SoID",
            "LoaiD": "LoaiID",
            "TenKhachHang": "TenKhachHang",
            "NgaySinh": "NgaySinh",
            "GioiTinh": "GioiTinh",
            "MaSoThue": "MaSoThue",
            "SoDienThoaiDangKyDichVu": "SoDienThoaiDangKyDichVu",
            "DiaChi": "DiaChi",
            "DiaChiKiemSoatTruyCap": "DiaChiKiemSoatTruyCap",
            "MaSoNhanDangThietBiDong": "MaSoNhanDangThietBiDiDong",
            "SoTaiKhoan": "SoTaiKhoan",
            "LoaiTaiKhoan": "LoaiTaiKhoan",
            "TrangThaiHoatDongTaiKhoan": "TrangThaiHoatDongTaiKhoan",
            "NgayMoTaiKhoan": "NgayMoTaiKhoan",
            "PhuongThucMoTaiKhoan": "PhuongThucMoTaiKhoan",
            "NgayXacThucTaiQuay": "NgayXacThucTaiQuay",
            "QuocTich": "QuocTich"
        }

        for record in data:
            converted_record = {}
            for db_field, simo_field in field_mapping.items():
                if db_field in record and record[db_field] is not None:
                    value = record[db_field]
                    
                    # Xử lý các trường số
                    if db_field in ["LoaiD", "GioiTinh", "LoaiTaiKhoan", 
                                  "TrangThaiHoatDongTaiKhoan", "PhuongThucMoTaiKhoan"]:
                        try:
                            converted_record[simo_field] = int(value)
                        except (ValueError, TypeError):
                            continue
                    
                    # Xử lý các trường cần padding
                    elif db_field in ["Cif", "Soid", "SoDienThoaiDangKyDichVu", "SoTaiKhoan"]:
                        formatted_value = SimoConverter.format_number_with_padding(value, simo_field)
                        if formatted_value:
                            converted_record[simo_field] = formatted_value
                    
                    # Các trường khác
                    else:
                        converted_record[simo_field] = str(value).strip()
            
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo_002(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_002"""
        payload = []
        field_mapping = {
            "Cif": "Cif",
            "SoTaiKhoan": "SoTaiKhoan",
            "TenKhachHang": "TenKhachHang",
            "TrangThaiHoatDongTaiKhoan": "TrangThaiHoatDongTaiKhoan",
            "NghiNgo": "NghiNgo",
            "GhiChu": "GhiChu"
        }

        for record in data:
            converted_record = {}
            for db_field, simo_field in field_mapping.items():
                if db_field in record:
                    value = record[db_field]
                    # Chắc chắn xuất NghiNgo dưới dạng int, mặc định là 0 nếu không có giá trị
                    if db_field == "NghiNgo":
                        try:
                            converted_record[simo_field] = int(value) if value not in [None, "", "None"] else 0
                        except (ValueError, TypeError):
                            converted_record[simo_field] = 0
                    # Xử lý các trường số nguyên khác
                    elif db_field == "TrangThaiHoatDongTaiKhoan":
                        try:
                            converted_record[simo_field] = int(value) if value not in [None, "", "None"] else None
                        except (ValueError, TypeError):
                            continue
                    # Xử lý các trường khác
                    elif value not in [None, "", "None"]:
                        converted_record[simo_field] = value
                    
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo_003(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_003"""
        payload = []
        field_mapping = {
            "Cif": "Cif",
            "SoTaiKhoan": "SoTaiKhoan",
            "TenKhachHang": "TenKhachHang",
            "TrangThaiHoatDongTaiKhoan": "TrangThaiHoatDongTaiKhoan",
            "NghiNgo": "NghiNgo"  # Thêm trường NghiNgo vào mapping
        }

        for record in data:
            converted_record = {}
            for db_field, simo_field in field_mapping.items():
                if db_field in record:
                    value = record[db_field]
                    # Chắc chắn xuất NghiNgo dưới dạng int, mặc định là 0 nếu không có giá trị
                    if db_field == "NghiNgo":
                        try:
                            converted_record[simo_field] = int(value) if value not in [None, "", "None"] else 0
                        except (ValueError, TypeError):
                            converted_record[simo_field] = 0
                    # Xử lý trường số nguyên khác
                    elif db_field == "TrangThaiHoatDongTaiKhoan":
                        try:
                            converted_record[simo_field] = int(value) if value not in [None, "", "None"] else None
                        except (ValueError, TypeError):
                            continue
                    # Xử lý các trường khác
                    elif value not in [None, "", "None"]:
                        converted_record[simo_field] = value
                    
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo_004(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_004"""
        payload = []
        fields = {
            "Cif": ('str', 36),
            "SoID": ('str', 15),
            "LoaiID": ('int', None),
            "TenKhachHang": ('str', 150),
            "NgaySinh": ('date', None),
            "GioiTinh": ('int', None),
            "MaSoThue": ('str', None),
            "SoDienThoaiDangKyDichVu": ('str', 15),
            "DiaChi": ('str', 300),
            "DiaChiKiemSoatTruyCap": ('str', 60),
            "MaSoNhanDangThietBiDiDong": ('str', 36),
            "SoTaiKhoan": ('str', None),
            "LoaiTaiKhoan": ('int', None),
            "TrangThaiHoatDongTaiKhoan": ('int', None),
            "NgayMoTaiKhoan": ('date', None),
            "PhuongThucMoTaiKhoan": ('int', None),
            "NgayXacThucTaiQuay": ('date', None),
            "GhiChu": ('str', 500),
            "QuocTich": ('str', 36)
        }
        
        for record in data:
            converted_record = {}
            for field, (field_type, max_length) in fields.items():
                # Xử lý các trường đặc biệt cần định dạng theo chuẩn VN
                if field in ["SoID", "SoDienThoaiDangKyDichVu", "Cif", "SoTaiKhoan"]:
                    value = record.get(field)
                    if value is not None:
                        formatted_value = SimoConverter.format_number_with_padding(value, field)
                        if formatted_value:
                            converted_record[field] = formatted_value
                else:
                    # Các trường khác xử lý như cũ
                    value = SimoConverter.get_default_value(record.get(field), field_type, max_length)
                    if value is not None:
                        converted_record[field] = value
                    
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo_011(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_011"""
        payload = []
        fields = {
            "Cif": ('str', 36),
            "SoID": ('str', 15),
            "LoaiID": ('int', None),
            "TenKhachHang": ('str', 150),
            "NgaySinh": ('date', None),
            "GioiTinh": ('int', None),
            "MaSoThue": ('str', None),
            "SoDienThoaiDangKyDichVu": ('str', 15),
            "DiaChi": ('str', 300),
            "DiaChiKiemSoatTruyCap": ('str', 60),
            "MaSoNhanDangThietBiDiDong": ('str', 36),
            "SoTaiKhoan": ('str', None),
            "LoaiTaiKhoan": ('int', None),
            "TrangThaiHoatDongTaiKhoan": ('int', None),
            "NgayMoTaiKhoan": ('date', None),
            "PhuongThucMoTaiKhoan": ('int', None),
            "NgayXacThucTaiQuay": ('date', None),
            "QuocTich": ('str', 36)
        }
        
        for record in data:
            converted_record = {}
            for field, (field_type, max_length) in fields.items():
                value = SimoConverter.get_default_value(record.get(field), field_type, max_length)
                if value is not None:
                    converted_record[field] = value
                    
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo_012(data):
        """Chuyển đổi dữ liệu sang định dạng SIMO_012"""
        payload = []
        fields = {
            "Cif": ('str', 36),
            "SoTaiKhoan": ('str', None),
            "TenKhachHang": ('str', 150),
            "TrangThaiHoatDongTaiKhoan": ('int', None),
            "NghiNgo": ('int', None),
            "GhiChu": ('str', 500)
        }
        
        for record in data:
            converted_record = {}
            for field, (field_type, max_length) in fields.items():
                value = SimoConverter.get_default_value(record.get(field), field_type, max_length)
                if value is not None:
                    converted_record[field] = value
                    
            if converted_record:
                payload.append(converted_record)
                
        return payload

    @staticmethod
    def convert_to_simo(data, service_type):
        """Chuyển đổi dữ liệu thành định dạng SIMO theo loại dịch vụ"""
        if service_type == "simo_001":
            return SimoConverter.convert_to_simo_001(data)
        elif service_type == "simo_002":
            return SimoConverter.convert_to_simo_002(data)
        elif service_type == "simo_003":
            return SimoConverter.convert_to_simo_003(data)
        elif service_type == "simo_004":
            return SimoConverter.convert_to_simo_004(data)
        elif service_type == "simo_011":
            return SimoConverter.convert_to_simo_011(data)
        elif service_type == "simo_012":
            return SimoConverter.convert_to_simo_012(data)
        else:
            raise ValueError(f"Không hỗ trợ định dạng SIMO: {service_type}")
            
    @staticmethod
    def convert_excel_to_json(file_path, service_type):
        """Chuyển đổi dữ liệu Excel sang định dạng JSON"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Lấy headers từ dòng đầu tiên
            headers = [str(cell.value).strip() for cell in ws[1]]
            
            # Đọc dữ liệu từ dòng thứ 2
            raw_data = []
            for row in ws.iter_rows(min_row=2):
                row_dict = {}
                for header, cell in zip(headers, row):
                    value = cell.value if cell.value is not None else ""
                    row_dict[header] = value
                raw_data.append(row_dict)

            # Chuyển đổi dữ liệu theo định dạng tương ứng
            return SimoConverter.convert_to_simo(raw_data, service_type)
                
        except Exception as e:
            logger.error(f"Lỗi khi chuyển đổi Excel sang JSON: {str(e)}")
            raise

    @staticmethod
    def validate_vietnam_phone_number(phone_number):
        """
        Kiểm tra tính hợp lệ của số điện thoại theo chuẩn Việt Nam
        Trả về tuple (is_valid, error_message)
        """
        if phone_number is None:
            return True, None
            
        # Làm sạch giá trị
        clean_value = str(phone_number).strip().replace(" ", "").replace("-", "")
        
        # Loại bỏ mã quốc gia +84 hoặc 84 nếu có
        if clean_value.startswith("+84"):
            clean_value = "0" + clean_value[3:]
        elif clean_value.startswith("84") and len(clean_value) >= 10:
            clean_value = "0" + clean_value[2:]
        
        # Kiểm tra độ dài
        if len(clean_value) != 10:
            return False, f"Số điện thoại '{phone_number}' không đúng 10 số theo chuẩn Việt Nam"
            
        # Kiểm tra số 0 ở đầu
        if not clean_value.startswith("0"):
            return False, f"Số điện thoại '{phone_number}' không bắt đầu bằng số 0 theo chuẩn Việt Nam"
            
        # Kiểm tra đầu số hợp lệ (03x, 05x, 07x, 08x, 09x)
        valid_prefixes = ["03", "05", "07", "08", "09"]
        if clean_value[:2] not in valid_prefixes:
            return False, f"Số điện thoại '{phone_number}' có đầu số {clean_value[:2]} không hợp lệ theo chuẩn Việt Nam"
            
        return True, None
        
    @staticmethod
    def validate_data_before_export(data, simo_type):
        """
        Kiểm tra dữ liệu trước khi xuất JSON
        Trả về tuple (is_valid, warning_messages)
        """
        is_valid = True
        warnings = []
        
        for i, record in enumerate(data):
            # Kiểm tra số điện thoại
            if "SoDienThoaiDangKyDichVu" in record:
                phone_valid, phone_error = SimoConverter.validate_vietnam_phone_number(record["SoDienThoaiDangKyDichVu"])
                if not phone_valid:
                    warnings.append(f"Bản ghi #{i+1}: {phone_error}")
                    is_valid = False
        
        return is_valid, warnings